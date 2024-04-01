import os
import ctypes
import openpyxl as xl

def message_box(title, text, style):
    return ctypes.windll.user32.MessageBoxW(0, text, title, style)

head = """:::SOURCE
:ACOM
1:ID:FCDRW,2.00:0;
2:DT:2024,04,01,08,28,27:1711920507:BKEEdtCtlDrw:6.60;
3:RC:2024,04,01,08,28,27:1711920507:BKEEdtCtlDrw:6.60;
::ACOM
:FDFL
::FDFL
:FHED
1:IT:::1600,1072:1;
2:CLT:HJ:1;
3:CLT:HG:2;
4:CLT:DJ:1;
5:CLT:DG:1;
::FHED
"""

PVI_template = """:FNRM
1:REV3K:1:1;
49:BKRV:2;
9:TPFX:1:;
2:ETAG:1:***TAG_NAME***;
3:EINS:1:PVI;
4:ETCM:1:***TAG_COMMENT***;
5:ETIM:1:2;
6:ELIM:1:4;
7:CHKN:1:***COUNT***;
10:ESCA:1:S;
11:SCCP:1:1:0;
12:MSBP:1:NO;
13:SREV:1:NO;
14:SQC!:1:0.5;
15:PIR!:1:1.00:AUTO;
16:OVPV:1:NO;
17:SSL!:1:1.000:0.000;
18:HLPM:1:0000;
19:UPPG:1:;
20:IVHS:1:1:1:2.0%;
21:EINP:1:SUBSYS;
22:SSI!:1:1.000:0.000:106.25:-6.25;
23:ESUM:1:NO:0%;
24:SMET:1:NO;
25:DBLA:1:NO;
26:DVNO:1:AUTO;
27:UPPL:1:;
28:ESCL:1:***SCALE***;
29:EUNT:1:***UNITS***;
30:FLTR:1:AUTO;
31:PVRL:1:NO;
32:SIOP:1:NO;
33:EALA:1:2;
34:INOP:1:HL;
35:BADS:1:0:0;
36:HHLL:1:HHLL;
37:INHL:1:HL;
38:HHLH:1:2.0%;
39:HHDS:1:0:0;
40:HIDS:1:0:0;
41:LODS:1:0:0;
42:LLDS:1:0:0;
43:INVC:1:NO;
44:ILCN:1:YES;
45:OUTL:1:LINEAR;
46:SUOP:1:PV;
47:SOAC:1:POSITION;
48:CNCT:1:***CONNECTION***;
8:GBLK:***COORDINATES***:E2;
::FNRM
"""

tail = "::::SOURCE"

x_ini, y_ini = 24, 24
x_delta, y_delta = 128, 88
x_max, y_max = 12, 9
n_max = x_max * y_max
max_blocks_per_sheet = 100 # cannot be more than 100

os.system('cls')

print("******  start of the script  *******")

excel_filename = 'input_data.xlsx'
try:
    wb = xl.load_workbook(excel_filename)
except Exception as e:
    print(f'Cannot open the excel file: {str(e)}, quitting...')
    message_box('Error', f'Cannot open the excel file: {str(e)}', 0)    
    quit()
print(f'{excel_filename} has been opened successfully')

sheet = wb['input data']

number_of_blocks = sheet.max_row - 1

for row in range(2, sheet.max_row + 1):
    if sheet.cell(row, 1).value in (None, ""):
        break
    
    COUNT = row - 1
    
    last_in_DR = COUNT % max_blocks_per_sheet == 0
    
    COUNT_IN_DR = max_blocks_per_sheet if last_in_DR else COUNT % max_blocks_per_sheet    

    if COUNT_IN_DR == 1:
        txt_out = head
    
    DR_count = row // max_blocks_per_sheet  
    index = COUNT_IN_DR - 1

    TAG_NAME    = sheet.cell(row, 3).value
    # TAG_NAME    = sheet.cell(row, 3).value + "P"
    TAG_COMMENT = sheet.cell(row, 4).value
    SL          = sheet.cell(row, 6).value
    SH          = sheet.cell(row, 7).value
    SCALE       = f"{SH}:{SL}"
    UNITS       = sheet.cell(row, 8).value
    CONNECTION  = f"IN:{TAG_NAME}.PV:I"
    
    temp = index - (index // x_max) * x_max
    x = x_ini + temp * x_delta    

    temp = index // x_max
    y = y_ini + temp * y_delta

    COORDINATES = f"{x}, {y}"

    PVI = PVI_template.replace("***COUNT***", str(COUNT_IN_DR))
    PVI = PVI.replace("***TAG_NAME***", TAG_NAME)
    PVI = PVI.replace("***TAG_COMMENT***", TAG_COMMENT)
    PVI = PVI.replace("***SCALE***", SCALE)
    PVI = PVI.replace("***UNITS***", UNITS)
    PVI = PVI.replace("***CONNECTION***", CONNECTION)
    PVI = PVI.replace("***COORDINATES***", COORDINATES)

    txt_out += PVI
    
    last_row = COUNT == number_of_blocks
    if last_in_DR or last_row:
        txt_out += tail
        if last_row:
            DR_count += 1
        with open(f"output/DR_IMPORT{DR_count}.txt", "w") as my_file:
            print(f'Saving the result into DR_IMPORT{DR_count}.txt...')
            my_file.write(str(txt_out))

wb.close()
print(f"*****   End of script     *****")