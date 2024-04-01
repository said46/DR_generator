import os
import openpyxl as xl

os.system('cls')

print ("*** start of the test ***")

excel_filename = 'input_data.xlsx'
try:
    wb = xl.load_workbook(excel_filename)
except Exception as e:
    print(f'Cannot open the excel file: {str(e)}, quitting...')
    quit()
print(f'{excel_filename} has been opened successfully')

sheet = wb['input data']

dict = {}
node_prev = "???"
number_of_blocks = sheet.max_row - 1
tag_counter = 0
COUNT_IN_DR = 0

for row in range(2, sheet.max_row + 1):
    if sheet.cell(row, 1).value in (None, ""):
        break  
    tag_counter = row - 1
    
    node = sheet.cell(row, 2).value
    if node != node_prev:
        if tag_counter != 1:
            print(f" end of {node_prev} with {COUNT_IN_DR=}") 
        node_prev = node
        COUNT_IN_DR = 0
        dict[node] = []
        print(f" start of {node}")
      
    COUNT_IN_DR += 1
    tag_name = sheet.cell(row, 3).value
    temp = (node, tag_name)
    print(temp)
    dict[node].append(temp)
    if tag_counter == number_of_blocks:
        print(f" end of {node_prev} with {COUNT_IN_DR=}") 

wb.close()    
print ("***  end of the test  ***")